"""
SIS Internal Marks — H57 Service.

Business rules:
  - Faculty must hold active PRIMARY or CO_FACULTY subject_assignment to create/edit
    components or enter marks for a course+section.
  - Component lifecycle: DRAFT → PUBLISHED → LOCKED.
    - DRAFT:      faculty can edit definition and enter marks freely.
    - PUBLISHED:  definition immutable; marks edits require edit_reason; students can see.
    - LOCKED:     no edits allowed; Dean/Admin can reopen (LOCKED → PUBLISHED) with reason.
  - Marks entry: marks_obtained > max_marks is rejected (422).
  - Edit reason is mandatory when:
      • component status is PUBLISHED and entry already has marks_obtained set.
  - Excel bulk import: two-phase preview/commit; partial success per row.
  - Readiness: advisory only — no autonomous grade, penalty, or rejection.
"""
from __future__ import annotations

import io
import logging
import uuid
from datetime import datetime, timezone
from decimal import Decimal
from typing import Optional
from uuid import UUID

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.audit_log.models import AuditEventType
from app.core.audit_log.service import AuditService
from app.modules.m11_sis.marks_models import ComponentStatus, SisMarksComponent
from app.modules.m11_sis.marks_repository import MarksRepository
from app.modules.m11_sis.marks_schemas import (
    BulkMarkEntryIn, BulkMarkEntryResult, ComponentCreateIn, ComponentOut,
    ComponentUpdateIn, MarkEntryOut, MarksImportCommitResult,
    MarksImportPreviewOut, MarksImportRowResult, MyMarksOut, ReopenComponentIn,
    SectionMarksReportOut, SectionReadinessOut, SingleMarkEditIn,
    StudentComponentView, StudentCourseMarks, StudentReadiness,
    SectionStudentMarks, FacultyCourseComponentSummary,
)
from app.modules.m_academics.class_roster import is_elective_course
from app.modules.m_academics.models import AcadSection, SubjectAssignment

logger = logging.getLogger("vidya.sis.marks")

DEFAULT_ATTENDANCE_THRESHOLD = 75.0


# ---------------------------------------------------------------------------
# Service error
# ---------------------------------------------------------------------------

class MarksServiceError(Exception):
    def __init__(self, code: str, message: str, status_code: int = 400) -> None:
        self.code = code
        self.message = message
        self.status_code = status_code
        super().__init__(message)


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def _component_out(row: dict) -> ComponentOut:
    return ComponentOut(
        id=row["id"],
        course_id=row["course_id"],
        course_code=row.get("course_code"),
        course_title=row.get("course_title"),
        section_id=row["section_id"],
        section_name=row.get("section_name"),
        semester_id=row["semester_id"],
        created_by=row["created_by"],
        component_type=row["component_type"],
        name=row["name"],
        max_marks=row["max_marks"],
        weightage=row.get("weightage"),
        due_date=row.get("due_date"),
        display_order=row.get("display_order"),
        status=row["status"],
        published_at=row.get("published_at"),
        locked_at=row.get("locked_at"),
        locked_by=row.get("locked_by"),
        reopened_by=row.get("reopened_by"),
        reopened_at=row.get("reopened_at"),
        reopen_reason=row.get("reopen_reason"),
        is_active=row["is_active"],
        entries_count=int(row.get("entries_count") or 0),
        filled_count=int(row.get("filled_count") or 0),
        created_at=row["created_at"],
        updated_at=row.get("updated_at"),
    )


def _entry_out(row: dict) -> MarkEntryOut:
    return MarkEntryOut(
        id=row["id"],
        component_id=row["component_id"],
        student_id=row["student_id"],
        student_name=row.get("student_name"),
        usn=row.get("usn"),
        marks_obtained=row.get("marks_obtained"),
        remarks=row.get("remarks"),
        marked_by=row.get("marked_by"),
        marked_at=row.get("marked_at"),
        edited_by=row.get("edited_by"),
        edited_at=row.get("edited_at"),
        edit_reason=row.get("edit_reason"),
        created_at=row["created_at"],
        updated_at=row.get("updated_at"),
    )


async def _assert_faculty_teaches(
    course_id: UUID,
    semester_id: UUID,
    faculty_id: UUID,
    db: AsyncSession,
) -> None:
    assignment = (
        await db.execute(
            select(SubjectAssignment)
            .where(SubjectAssignment.course_id       == course_id)
            .where(SubjectAssignment.semester_id     == semester_id)
            .where(SubjectAssignment.faculty_user_id == faculty_id)
            .where(SubjectAssignment.is_active.is_(True))
            .where(SubjectAssignment.role_in_course.in_(["PRIMARY", "CO_FACULTY"]))
        )
    ).scalar_one_or_none()

    if assignment is None:
        raise MarksServiceError(
            "NOT_ASSIGNED",
            "You must be an active PRIMARY or CO_FACULTY for this course to manage marks.",
            403,
        )


async def _verify_faculty_assignment(
    course_id: UUID,
    section_id: UUID,
    faculty_id: UUID,
    db: AsyncSession,
) -> AcadSection:
    section = (
        await db.execute(select(AcadSection).where(AcadSection.id == section_id))
    ).scalar_one_or_none()
    if section is None:
        raise MarksServiceError("SECTION_NOT_FOUND", "Section not found.", 404)
    if not section.is_active:
        raise MarksServiceError("SECTION_INACTIVE", "Section is inactive.", 400)

    await _assert_faculty_teaches(course_id, section.semester_id, faculty_id, db)
    return section


async def _resolve_component_class(
    body: ComponentCreateIn,
    faculty_id: UUID,
    db: AsyncSession,
) -> tuple[Optional[UUID], UUID]:
    """Which class this component belongs to: `(section_id, semester_id)`.

    `section_id is None` marks an elective group — one component covering every
    student who chose the elective, not one per contributing section. Mirrors
    `attendance_service._resolve_session_class` so the two rosters agree.
    """
    if await is_elective_course(body.course_id, db):
        if body.semester_id is None:
            raise MarksServiceError(
                "SEMESTER_REQUIRED",
                "This is an elective. Pass semester_id — its class is every student "
                "who chose it this semester, not a section.",
                422,
            )
        await _assert_faculty_teaches(body.course_id, body.semester_id, faculty_id, db)
        return None, body.semester_id

    if body.section_id is None:
        raise MarksServiceError(
            "SECTION_REQUIRED", "This course is taught per section. Pass section_id.", 422,
        )
    section = await _verify_faculty_assignment(body.course_id, body.section_id, faculty_id, db)
    return body.section_id, section.semester_id


async def _notify_safe(*, notification_type, recipient_user_id, recipient_email, title, body, entity_id, db):
    try:
        from app.core.notifications.service import NotificationService
        await NotificationService.send(
            notification_type,
            recipient_user_id=recipient_user_id,
            recipient_email=recipient_email,
            title=title,
            body=body,
            entity_type="marks_component",
            entity_id=entity_id,
            db=db,
        )
    except Exception:
        logger.warning("marks notification failed type=%s recipient=%s", notification_type, recipient_user_id, exc_info=True)


# ---------------------------------------------------------------------------
# MarksService
# ---------------------------------------------------------------------------

class MarksService:

    # ------------------------------------------------------------------
    # Components — CRUD
    # ------------------------------------------------------------------

    @staticmethod
    async def create_component(
        body: ComponentCreateIn,
        actor_id: UUID,
        actor_role: str,
        tenant_id: Optional[UUID],
        schema_name: Optional[str],
        db: AsyncSession,
    ) -> ComponentOut:
        section_id, semester_id = await _resolve_component_class(body, actor_id, db)

        component = SisMarksComponent(
            id=uuid.uuid4(),
            course_id=body.course_id,
            section_id=section_id,
            semester_id=semester_id,
            created_by=actor_id,
            component_type=body.component_type,
            name=body.name,
            max_marks=body.max_marks,
            weightage=body.weightage,
            due_date=body.due_date,
            display_order=body.display_order,
            status=ComponentStatus.DRAFT,
        )
        component = await MarksRepository.create_component(component, db)
        await db.commit()

        await AuditService.log(
            AuditEventType.INTERNAL_MARKS_COMPONENT_CREATED,
            actor_user_id=actor_id,
            actor_role=actor_role,
            tenant_id=tenant_id,
            schema_name=schema_name,
            target_entity="SisMarksComponent",
            target_id=str(component.id),
            metadata={
                "name": body.name,
                "component_type": body.component_type,
                "max_marks": float(body.max_marks),
                "course_id": str(body.course_id),
                "section_id": str(section_id) if section_id else None,
                "semester_id": str(semester_id),
                "is_elective_group": section_id is None,
            },
        )

        rows = await MarksRepository.list_components_enriched(
            course_id=body.course_id, section_id=section_id, db=db
        )
        row = next((r for r in rows if r["id"] == component.id), None)
        if row is None:
            raise MarksServiceError("INTERNAL", "Component created but not retrievable.", 500)
        return _component_out(row)

    @staticmethod
    async def list_components(
        *,
        course_id:   Optional[UUID] = None,
        section_id:  Optional[UUID] = None,
        semester_id: Optional[UUID] = None,
        status:      Optional[str]  = None,
        actor_id:    UUID,
        actor_role:  str,
        db: AsyncSession,
    ) -> list[ComponentOut]:
        faculty_filter = actor_id if actor_role == "FACULTY" else None
        rows = await MarksRepository.list_components_enriched(
            course_id=course_id,
            section_id=section_id,
            semester_id=semester_id,
            status=status,
            faculty_id=faculty_filter,
            db=db,
        )
        return [_component_out(r) for r in rows]

    @staticmethod
    async def get_component(component_id: UUID, actor_id: UUID, actor_role: str, db: AsyncSession) -> ComponentOut:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if actor_role == "FACULTY" and component.created_by != actor_id:
            raise MarksServiceError("FORBIDDEN", "You do not own this component.", 403)

        rows = await MarksRepository.list_components_enriched(
            course_id=component.course_id, section_id=component.section_id, db=db
        )
        row = next((r for r in rows if r["id"] == component_id), None)
        if row is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        return _component_out(row)

    @staticmethod
    async def update_component(
        component_id: UUID,
        body: ComponentUpdateIn,
        actor_id: UUID,
        actor_role: str,
        tenant_id: Optional[UUID],
        schema_name: Optional[str],
        db: AsyncSession,
    ) -> ComponentOut:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if actor_role == "FACULTY" and component.created_by != actor_id:
            raise MarksServiceError("FORBIDDEN", "You do not own this component.", 403)
        if component.status != ComponentStatus.DRAFT:
            raise MarksServiceError(
                "NOT_DRAFT",
                "Component definition can only be edited while in DRAFT status.",
                409,
            )

        now = datetime.now(timezone.utc)
        if body.component_type is not None:
            component.component_type = body.component_type
        if body.name is not None:
            component.name = body.name
        if body.max_marks is not None:
            component.max_marks = body.max_marks
        if body.weightage is not None:
            component.weightage = body.weightage
        if body.due_date is not None:
            component.due_date = body.due_date
        if body.display_order is not None:
            component.display_order = body.display_order
        component.updated_at = now

        await MarksRepository.update_component(component, db)
        await db.commit()

        await AuditService.log(
            AuditEventType.INTERNAL_MARKS_COMPONENT_UPDATED,
            actor_user_id=actor_id,
            actor_role=actor_role,
            tenant_id=tenant_id,
            schema_name=schema_name,
            target_entity="SisMarksComponent",
            target_id=str(component_id),
        )

        rows = await MarksRepository.list_components_enriched(
            course_id=component.course_id, section_id=component.section_id, db=db
        )
        row = next((r for r in rows if r["id"] == component_id), None)
        if row is None:
            raise MarksServiceError("INTERNAL", "Component not found after update.", 500)
        return _component_out(row)

    @staticmethod
    async def delete_component(
        component_id: UUID,
        actor_id: UUID,
        actor_role: str,
        tenant_id: Optional[UUID],
        schema_name: Optional[str],
        db: AsyncSession,
    ) -> None:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if actor_role == "FACULTY" and component.created_by != actor_id:
            raise MarksServiceError("FORBIDDEN", "You do not own this component.", 403)
        if component.status != ComponentStatus.DRAFT:
            raise MarksServiceError(
                "NOT_DRAFT",
                "Only DRAFT components can be deleted.",
                409,
            )

        component.is_active = False
        component.updated_at = datetime.now(timezone.utc)
        await db.flush()
        await db.commit()

        await AuditService.log(
            AuditEventType.INTERNAL_MARKS_COMPONENT_UPDATED,
            actor_user_id=actor_id,
            actor_role=actor_role,
            tenant_id=tenant_id,
            schema_name=schema_name,
            target_entity="SisMarksComponent",
            target_id=str(component_id),
            metadata={"action": "soft_delete"},
        )

    # ------------------------------------------------------------------
    # Lifecycle transitions
    # ------------------------------------------------------------------

    @staticmethod
    async def publish_component(
        component_id: UUID,
        actor_id: UUID,
        actor_role: str,
        tenant_id: Optional[UUID],
        schema_name: Optional[str],
        db: AsyncSession,
    ) -> ComponentOut:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if actor_role == "FACULTY" and component.created_by != actor_id:
            raise MarksServiceError("FORBIDDEN", "You do not own this component.", 403)
        if component.status != ComponentStatus.DRAFT:
            raise MarksServiceError("NOT_DRAFT", "Only DRAFT components can be published.", 409)

        now = datetime.now(timezone.utc)
        component.status       = ComponentStatus.PUBLISHED
        component.published_at = now
        component.updated_at   = now
        await db.flush()
        await db.commit()

        await AuditService.log(
            AuditEventType.INTERNAL_MARKS_SUBMITTED,   # reuse H35 stub
            actor_user_id=actor_id,
            actor_role=actor_role,
            tenant_id=tenant_id,
            schema_name=schema_name,
            target_entity="SisMarksComponent",
            target_id=str(component_id),
            metadata={
                "name": component.name,
                "course_id": str(component.course_id),
                "section_id": str(component.section_id),
            },
        )

        # Notify all enrolled students
        await MarksService._notify_students_on_publish(component, db)

        rows = await MarksRepository.list_components_enriched(
            course_id=component.course_id, section_id=component.section_id, db=db
        )
        row = next((r for r in rows if r["id"] == component_id), None)
        if row is None:
            raise MarksServiceError("INTERNAL", "Component not found after publish.", 500)
        return _component_out(row)

    @staticmethod
    async def _notify_students_on_publish(component: SisMarksComponent, db: AsyncSession) -> None:
        try:
            from app.core.notifications.models import NotificationType
            students = await MarksRepository.get_class_students(component, db)
            course_row = (await db.execute(
                text("SELECT code, title FROM courses WHERE id = :id"),
                {"id": str(component.course_id)},
            )).one_or_none()
            course_code = course_row[0] if course_row else "your course"
            for s in students:
                await _notify_safe(
                    notification_type=NotificationType.INTERNAL_MARKS_PUBLISHED,
                    recipient_user_id=UUID(str(s["student_id"])),
                    recipient_email=s.get("email"),
                    title=f"Marks Published — {course_code}",
                    body=(
                        f"Your {component.name} marks for {course_code} are now available. "
                        "Log in to view your marks."
                    ),
                    entity_id=str(component.id),
                    db=db,
                )
        except Exception:
            logger.warning("marks publish notify failed component=%s", component.id, exc_info=True)

    @staticmethod
    async def lock_component(
        component_id: UUID,
        actor_id: UUID,
        actor_role: str,
        tenant_id: Optional[UUID],
        schema_name: Optional[str],
        db: AsyncSession,
    ) -> ComponentOut:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if component.status != ComponentStatus.PUBLISHED:
            raise MarksServiceError("NOT_PUBLISHED", "Only PUBLISHED components can be locked.", 409)

        now = datetime.now(timezone.utc)
        component.status    = ComponentStatus.LOCKED
        component.locked_at = now
        component.locked_by = actor_id
        component.updated_at = now
        await db.flush()
        await db.commit()

        await AuditService.log(
            AuditEventType.INTERNAL_MARKS_LOCKED,   # reuse H35 stub
            actor_user_id=actor_id,
            actor_role=actor_role,
            tenant_id=tenant_id,
            schema_name=schema_name,
            target_entity="SisMarksComponent",
            target_id=str(component_id),
            metadata={
                "name": component.name,
                "course_id": str(component.course_id),
                "section_id": str(component.section_id),
                "locked_by": str(actor_id),
            },
        )

        rows = await MarksRepository.list_components_enriched(
            course_id=component.course_id, section_id=component.section_id, db=db
        )
        row = next((r for r in rows if r["id"] == component_id), None)
        if row is None:
            raise MarksServiceError("INTERNAL", "Component not found after lock.", 500)
        return _component_out(row)

    @staticmethod
    async def reopen_component(
        component_id: UUID,
        body: ReopenComponentIn,
        actor_id: UUID,
        actor_role: str,
        tenant_id: Optional[UUID],
        schema_name: Optional[str],
        db: AsyncSession,
    ) -> ComponentOut:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if component.status != ComponentStatus.LOCKED:
            raise MarksServiceError("NOT_LOCKED", "Only LOCKED components can be reopened.", 409)

        now = datetime.now(timezone.utc)
        component.status       = ComponentStatus.PUBLISHED
        component.reopened_by  = actor_id
        component.reopened_at  = now
        component.reopen_reason = body.reason
        component.updated_at   = now
        await db.flush()
        await db.commit()

        await AuditService.log(
            AuditEventType.INTERNAL_MARKS_COMPONENT_REOPENED,
            actor_user_id=actor_id,
            actor_role=actor_role,
            tenant_id=tenant_id,
            schema_name=schema_name,
            target_entity="SisMarksComponent",
            target_id=str(component_id),
            metadata={
                "reason": body.reason,
                "reopened_by": str(actor_id),
            },
        )

        rows = await MarksRepository.list_components_enriched(
            course_id=component.course_id, section_id=component.section_id, db=db
        )
        row = next((r for r in rows if r["id"] == component_id), None)
        if row is None:
            raise MarksServiceError("INTERNAL", "Component not found after reopen.", 500)
        return _component_out(row)

    # ------------------------------------------------------------------
    # Mark entries
    # ------------------------------------------------------------------

    @staticmethod
    async def get_entries(
        component_id: UUID,
        actor_id: UUID,
        actor_role: str,
        db: AsyncSession,
    ) -> list[MarkEntryOut]:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if actor_role == "FACULTY" and component.created_by != actor_id:
            raise MarksServiceError("FORBIDDEN", "You do not own this component.", 403)
        rows = await MarksRepository.get_entries_enriched(component_id, db)
        return [_entry_out(r) for r in rows]

    @staticmethod
    async def bulk_enter_marks(
        component_id: UUID,
        body: BulkMarkEntryIn,
        actor_id: UUID,
        actor_role: str,
        tenant_id: Optional[UUID],
        schema_name: Optional[str],
        db: AsyncSession,
    ) -> BulkMarkEntryResult:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if actor_role == "FACULTY" and component.created_by != actor_id:
            raise MarksServiceError("FORBIDDEN", "You do not own this component.", 403)
        if component.status == ComponentStatus.LOCKED:
            raise MarksServiceError("MARKS_LOCKED", "Marks are locked and cannot be edited.", 409)

        # Validate all marks <= max_marks
        for item in body.entries:
            if item.marks_obtained is not None and item.marks_obtained > component.max_marks:
                raise MarksServiceError(
                    "MARKS_EXCEEDED",
                    f"marks_obtained ({item.marks_obtained}) exceeds max_marks ({component.max_marks}).",
                    422,
                )

        # Check if any entries are already saved (determines edit_reason requirement)
        existing_count_row = (await db.execute(
            text("SELECT COUNT(*) FROM sis_marks_entries WHERE component_id = :cid AND marked_by IS NOT NULL"),
            {"cid": str(component_id)},
        )).scalar()
        has_existing = int(existing_count_row or 0) > 0

        if has_existing and component.status == ComponentStatus.PUBLISHED and not body.edit_reason:
            raise MarksServiceError(
                "EDIT_REASON_REQUIRED",
                "edit_reason is required when editing marks on a PUBLISHED component.",
                400,
            )

        entries = [
            {
                "student_id": item.student_id,
                "marks_obtained": item.marks_obtained,
                "remarks": item.remarks,
            }
            for item in body.entries
        ]

        saved, first_marks, edits = await MarksRepository.bulk_upsert_entries(
            component_id=component_id,
            entries=entries,
            actor_id=actor_id,
            edit_reason=body.edit_reason,
            is_first_bulk=not has_existing,
            db=db,
        )
        await db.commit()

        event = AuditEventType.INTERNAL_MARKS_ENTERED if first_marks > 0 else AuditEventType.INTERNAL_MARKS_EDITED
        await AuditService.log(
            event,
            actor_user_id=actor_id,
            actor_role=actor_role,
            tenant_id=tenant_id,
            schema_name=schema_name,
            target_entity="SisMarksComponent",
            target_id=str(component_id),
            metadata={"saved": saved, "first_marks": first_marks, "edits": edits},
        )

        return BulkMarkEntryResult(
            component_id=component_id,
            saved=saved,
            first_marks=first_marks,
            edits=edits,
        )

    @staticmethod
    async def edit_single_entry(
        component_id: UUID,
        entry_id: UUID,
        body: SingleMarkEditIn,
        actor_id: UUID,
        actor_role: str,
        tenant_id: Optional[UUID],
        schema_name: Optional[str],
        db: AsyncSession,
    ) -> MarkEntryOut:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if actor_role == "FACULTY" and component.created_by != actor_id:
            raise MarksServiceError("FORBIDDEN", "You do not own this component.", 403)
        if component.status == ComponentStatus.LOCKED:
            raise MarksServiceError("MARKS_LOCKED", "Marks are locked and cannot be edited.", 409)

        if body.marks_obtained is not None and body.marks_obtained > component.max_marks:
            raise MarksServiceError(
                "MARKS_EXCEEDED",
                f"marks_obtained ({body.marks_obtained}) exceeds max_marks ({component.max_marks}).",
                422,
            )

        entry = await MarksRepository.get_entry_by_id(entry_id, db)
        if entry is None or entry.component_id != component_id:
            raise MarksServiceError("ENTRY_NOT_FOUND", "Mark entry not found.", 404)

        now = datetime.now(timezone.utc)
        is_first = entry.marked_by is None

        if not is_first and component.status == ComponentStatus.PUBLISHED and not body.edit_reason:
            raise MarksServiceError(
                "EDIT_REASON_REQUIRED",
                "edit_reason is required when editing marks on a PUBLISHED component.",
                400,
            )

        if is_first:
            entry.marked_by = actor_id
            entry.marked_at = now
        else:
            entry.edited_by   = actor_id
            entry.edited_at   = now
            entry.edit_reason = body.edit_reason

        if body.marks_obtained is not None:
            entry.marks_obtained = body.marks_obtained
        if body.remarks is not None:
            entry.remarks = body.remarks
        entry.updated_at = now

        await MarksRepository.update_entry(entry, db)
        await db.commit()

        await AuditService.log(
            AuditEventType.INTERNAL_MARKS_EDITED,
            actor_user_id=actor_id,
            actor_role=actor_role,
            tenant_id=tenant_id,
            schema_name=schema_name,
            target_entity="SisMarksEntry",
            target_id=str(entry_id),
            metadata={
                "component_id": str(component_id),
                "student_id": str(entry.student_id),
                "marks_obtained": float(body.marks_obtained) if body.marks_obtained is not None else None,
                "edit_reason": body.edit_reason,
            },
        )

        rows = await MarksRepository.get_entries_enriched(component_id, db)
        row = next((r for r in rows if r["id"] == entry_id), None)
        if row is None:
            raise MarksServiceError("INTERNAL", "Entry not found after update.", 500)
        return _entry_out(row)

    # ------------------------------------------------------------------
    # Excel import — two-phase
    # ------------------------------------------------------------------

    @staticmethod
    async def import_preview(
        component_id: UUID,
        content: bytes,
        filename: str,
        actor_id: UUID,
        actor_role: str,
        db: AsyncSession,
    ) -> MarksImportPreviewOut:
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)
        if actor_role == "FACULTY" and component.created_by != actor_id:
            raise MarksServiceError("FORBIDDEN", "You do not own this component.", 403)
        if component.status == ComponentStatus.LOCKED:
            raise MarksServiceError("MARKS_LOCKED", "Marks are locked.", 409)

        from app.core.onboarding.service import OnboardingService
        raw_rows, parse_err = OnboardingService._parse_file(
            content, filename, required_cols=["email", "marks"]
        )
        if parse_err:
            return MarksImportPreviewOut(
                component_id=component_id,
                component_name=component.name,
                max_marks=component.max_marks,
                total_rows=0,
                valid_rows=0,
                invalid_rows=0,
                rows=[MarksImportRowResult(row_number=0, is_valid=False, errors=[parse_err])],
            )

        # Build email → student lookup
        enrolled = await MarksRepository.get_class_students(component, db)
        email_map = {s["email"].lower(): s for s in enrolled}

        rows: list[MarksImportRowResult] = []
        for i, raw in enumerate(raw_rows):
            errors: list[str] = []
            email = (raw.get("email") or "").strip().lower()
            marks_raw = (raw.get("marks") or "").strip()
            remarks = (raw.get("remarks") or "").strip() or None

            marks_val: Decimal | None = None
            if marks_raw:
                try:
                    marks_val = Decimal(marks_raw)
                    if marks_val < 0:
                        errors.append("marks cannot be negative")
                    elif marks_val > component.max_marks:
                        errors.append(f"marks ({marks_val}) exceed max_marks ({component.max_marks})")
                except Exception:
                    errors.append(f"marks '{marks_raw}' is not a valid number")

            student = email_map.get(email)
            if not student:
                errors.append("student not found in this section's enrollment")

            rows.append(MarksImportRowResult(
                row_number=i + 1,
                email=email,
                student_name=student["full_name"] if student else None,
                student_id=UUID(str(student["student_id"])) if student else None,
                usn=student.get("usn") if student else None,
                marks_obtained=marks_val if not errors else None,
                remarks=remarks,
                is_valid=len(errors) == 0,
                errors=errors,
            ))

        valid = sum(1 for r in rows if r.is_valid)
        return MarksImportPreviewOut(
            component_id=component_id,
            component_name=component.name,
            max_marks=component.max_marks,
            total_rows=len(rows),
            valid_rows=valid,
            invalid_rows=len(rows) - valid,
            rows=rows,
        )

    @staticmethod
    async def import_commit(
        component_id: UUID,
        content: bytes,
        filename: str,
        edit_reason: Optional[str],
        actor_id: UUID,
        actor_role: str,
        tenant_id: Optional[UUID],
        schema_name: Optional[str],
        db: AsyncSession,
    ) -> MarksImportCommitResult:
        preview = await MarksService.import_preview(
            component_id, content, filename, actor_id, actor_role, db
        )

        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)

        # Check if edit_reason needed
        existing_count_row = (await db.execute(
            text("SELECT COUNT(*) FROM sis_marks_entries WHERE component_id = :cid AND marked_by IS NOT NULL"),
            {"cid": str(component_id)},
        )).scalar()
        has_existing = int(existing_count_row or 0) > 0
        if has_existing and component.status == ComponentStatus.PUBLISHED and not edit_reason:
            raise MarksServiceError(
                "EDIT_REASON_REQUIRED",
                "edit_reason is required when importing marks on a PUBLISHED component.",
                400,
            )

        saved = 0
        row_errors: list[str] = []
        entries_to_upsert = []
        for row in preview.rows:
            if not row.is_valid or row.student_id is None:
                if row.errors:
                    row_errors.append(f"Row {row.row_number}: {'; '.join(row.errors)}")
                continue
            entries_to_upsert.append({
                "student_id": row.student_id,
                "marks_obtained": row.marks_obtained,
                "remarks": row.remarks,
            })

        if entries_to_upsert:
            saved_cnt, _, _ = await MarksRepository.bulk_upsert_entries(
                component_id=component_id,
                entries=entries_to_upsert,
                actor_id=actor_id,
                edit_reason=edit_reason,
                is_first_bulk=not has_existing,
                db=db,
            )
            saved = saved_cnt
            await db.commit()

            await AuditService.log(
                AuditEventType.INTERNAL_MARKS_ENTERED,
                actor_user_id=actor_id,
                actor_role=actor_role,
                tenant_id=tenant_id,
                schema_name=schema_name,
                target_entity="SisMarksComponent",
                target_id=str(component_id),
                metadata={"source": "excel_import", "saved": saved},
            )

        return MarksImportCommitResult(
            component_id=component_id,
            total=preview.total_rows,
            saved=saved,
            skipped=preview.total_rows - saved,
            errors=row_errors,
        )

    # ------------------------------------------------------------------
    # Excel template download
    # ------------------------------------------------------------------

    @staticmethod
    async def get_excel_template(component_id: UUID, db: AsyncSession) -> tuple[bytes, str]:
        """
        Generate XLSX with headers + enrolled student list pre-filled.
        Returns (bytes, filename).
        """
        component = await MarksRepository.get_component_by_id(component_id, db)
        if component is None:
            raise MarksServiceError("COMPONENT_NOT_FOUND", "Component not found.", 404)

        rows = await MarksRepository.get_enrolled_with_entries(component, db)

        import openpyxl
        from openpyxl.styles import Font, PatternFill
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Marks"

        headers = ["email", "student_name", "usn", f"marks (max: {component.max_marks})", "remarks"]
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            # rename the column header key for parsing
            if h.startswith("marks"):
                ws.cell(row=1, column=col, value="marks")

        for r_idx, s in enumerate(rows, 2):
            ws.cell(row=r_idx, column=1, value=s["email"])
            ws.cell(row=r_idx, column=2, value=s["full_name"])
            ws.cell(row=r_idx, column=3, value=s.get("usn") or "")
            existing_marks = s.get("marks_obtained")
            ws.cell(row=r_idx, column=4, value=float(existing_marks) if existing_marks is not None else None)
            ws.cell(row=r_idx, column=5, value=s.get("remarks") or "")

        # Column widths
        for col, width in zip("ABCDE", [32, 30, 15, 12, 30]):
            ws.column_dimensions[col].width = width

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        safe_name = component.name.replace(" ", "_").replace("/", "-")
        filename = f"marks_{safe_name}.xlsx"
        return buf.getvalue(), filename

    # ------------------------------------------------------------------
    # Analytics — section report
    # ------------------------------------------------------------------

    @staticmethod
    async def get_section_marks_report(
        section_id: UUID,
        course_id: Optional[UUID],
        db: AsyncSession,
    ) -> SectionMarksReportOut:
        # Fetch components
        comp_rows = await MarksRepository.list_components_enriched(
            section_id=section_id, course_id=course_id, db=db
        )
        components = [_component_out(r) for r in comp_rows]

        # Section + course metadata
        ctx = (await db.execute(
            text("""
                SELECT s.name AS section_name,
                       c.id AS course_id, c.code AS course_code, c.title AS course_title
                FROM acad_sections s
                CROSS JOIN courses c
                WHERE s.id = :sid
                LIMIT 1
            """),
            {"sid": str(section_id)},
        )).one_or_none()

        raw = await MarksRepository.get_section_marks_raw(section_id, course_id, db)

        # Aggregate per student
        student_map: dict[str, dict] = {}
        component_max: dict[str, Decimal] = {str(r["id"]): r["max_marks"] for r in comp_rows}

        for row in raw:
            sid_str = str(row["student_id"])
            if sid_str not in student_map:
                student_map[sid_str] = {
                    "student_id":   UUID(sid_str),
                    "student_name": row["student_name"],
                    "usn":          row.get("usn"),
                    "total_marks":  Decimal(0),
                    "total_max":    Decimal(0),
                    "filled":       0,
                    "total_comp":   len(component_max),
                }
            if row["marks_obtained"] is not None:
                student_map[sid_str]["total_marks"] += Decimal(str(row["marks_obtained"]))
                student_map[sid_str]["filled"] += 1
            cid_str = str(row.get("component_id", ""))
            if cid_str in component_max:
                student_map[sid_str]["total_max"] += component_max[cid_str]

        students: list[SectionStudentMarks] = []
        for sd in student_map.values():
            fill_pct = (
                round(sd["filled"] / sd["total_comp"] * 100, 1)
                if sd["total_comp"] > 0 else None
            )
            students.append(SectionStudentMarks(
                student_id=sd["student_id"],
                student_name=sd["student_name"],
                usn=sd["usn"],
                total_marks=sd["total_marks"] if sd["filled"] > 0 else None,
                total_max=sd["total_max"] if sd["total_comp"] > 0 else None,
                fill_pct=fill_pct,
            ))

        return SectionMarksReportOut(
            section_id=section_id,
            section_name=ctx[0] if ctx else "",
            course_id=UUID(str(ctx[1])) if ctx and course_id else (course_id or uuid.uuid4()),
            course_code=ctx[2] if ctx else "",
            course_title=ctx[3] if ctx else "",
            components=components,
            students=students,
        )

    # ------------------------------------------------------------------
    # Analytics — student self-view
    # ------------------------------------------------------------------

    @staticmethod
    async def get_my_marks(student_id: UUID, db: AsyncSession) -> MyMarksOut:
        from app.core.auth.models import User
        user = (await db.execute(select(User).where(User.id == student_id))).scalar_one_or_none()
        if user is None:
            raise MarksServiceError("STUDENT_NOT_FOUND", "Student not found.", 404)

        profile_row = (await db.execute(
            text("SELECT usn FROM sis_student_profiles WHERE user_id = :uid"),
            {"uid": str(student_id)},
        )).one_or_none()
        usn = profile_row[0] if profile_row else None

        raw = await MarksRepository.get_student_marks_raw(student_id, db)

        # Group by course+section
        course_map: dict[tuple, dict] = {}
        for row in raw:
            key = (str(row["course_id"]), str(row["section_id"]))
            if key not in course_map:
                course_map[key] = {
                    "course_id":   UUID(str(row["course_id"])),
                    "course_code": row["course_code"],
                    "course_title": row["course_title"],
                    "section_id":  UUID(str(row["section_id"])),
                    "section_name": row["section_name"],
                    "components":  [],
                    "total_obtained": Decimal(0),
                    "total_max":      Decimal(0),
                    "has_marks":      False,
                }
            entry = course_map[key]
            entry["components"].append(StudentComponentView(
                component_id=UUID(str(row["component_id"])),
                component_type=row["component_type"],
                name=row["component_name"],
                max_marks=Decimal(str(row["max_marks"])),
                weightage=Decimal(str(row["weightage"])) if row.get("weightage") else None,
                due_date=row.get("due_date"),
                display_order=row.get("display_order"),
                status=row["component_status"],
                marks_obtained=Decimal(str(row["marks_obtained"])) if row.get("marks_obtained") is not None else None,
                remarks=row.get("remarks"),
            ))
            entry["total_max"] += Decimal(str(row["max_marks"]))
            if row.get("marks_obtained") is not None:
                entry["total_obtained"] += Decimal(str(row["marks_obtained"]))
                entry["has_marks"] = True

        courses: list[StudentCourseMarks] = [
            StudentCourseMarks(
                course_id=cd["course_id"],
                course_code=cd["course_code"],
                course_title=cd["course_title"],
                section_id=cd["section_id"],
                section_name=cd["section_name"],
                total_marks_obtained=cd["total_obtained"] if cd["has_marks"] else None,
                total_max_marks=cd["total_max"] if cd["total_max"] > 0 else None,
                components=cd["components"],
            )
            for cd in course_map.values()
        ]

        return MyMarksOut(
            student_id=student_id,
            student_name=user.full_name,
            usn=usn,
            courses=courses,
        )

    # ------------------------------------------------------------------
    # Faculty course overview
    # ------------------------------------------------------------------

    @staticmethod
    async def get_my_courses_overview(faculty_id: UUID, db: AsyncSession) -> list[FacultyCourseComponentSummary]:
        rows = await MarksRepository.list_components_enriched(faculty_id=faculty_id, db=db)

        group: dict[tuple, dict] = {}
        for row in rows:
            key = (str(row["course_id"]), str(row["section_id"]))
            if key not in group:
                enrolled_count_row = (await db.execute(
                    text("SELECT COUNT(*) FROM acad_enrollments WHERE section_id = :sid AND is_active = true"),
                    {"sid": str(row["section_id"])},
                )).scalar()
                group[key] = {
                    "course_id":      UUID(str(row["course_id"])),
                    "course_code":    row.get("course_code") or "",
                    "course_title":   row.get("course_title") or "",
                    "section_id":     UUID(str(row["section_id"])),
                    "section_name":   row.get("section_name") or "",
                    "total_students": int(enrolled_count_row or 0),
                    "components":     [],
                }
            group[key]["components"].append(_component_out(row))

        return [
            FacultyCourseComponentSummary(
                course_id=v["course_id"],
                course_code=v["course_code"],
                course_title=v["course_title"],
                section_id=v["section_id"],
                section_name=v["section_name"],
                total_students=v["total_students"],
                components=v["components"],
            )
            for v in group.values()
        ]

    # ------------------------------------------------------------------
    # Readiness indicators
    # ------------------------------------------------------------------

    @staticmethod
    async def get_section_readiness(
        section_id: UUID,
        attendance_threshold: float,
        db: AsyncSession,
    ) -> SectionReadinessOut:
        raw = await MarksRepository.get_section_readiness_raw(section_id, attendance_threshold, db)

        section_row = (await db.execute(
            text("SELECT name FROM acad_sections WHERE id = :sid"), {"sid": str(section_id)}
        )).one_or_none()
        section_name = section_row[0] if section_row else ""

        students: list[StudentReadiness] = []
        att_ready = marks_ready = fully = 0
        for row in raw:
            att_pct = float(row["attendance_pct"]) if row["attendance_pct"] is not None else None
            total_locked = int(row["total_locked"] or 0)
            filled = int(row["marks_filled"] or 0)

            a_ready = att_pct is not None and att_pct >= attendance_threshold
            m_ready = total_locked > 0 and filled >= total_locked

            marks_fill_pct = round(filled / total_locked * 100, 1) if total_locked > 0 else None
            notes: list[str] = []
            if att_pct is None:
                notes.append("No finalized attendance sessions found.")
            elif not a_ready:
                notes.append(f"Attendance {att_pct:.1f}% is below threshold {attendance_threshold:.0f}%.")
            if total_locked == 0:
                notes.append("No LOCKED internal marks components for this section.")
            elif not m_ready:
                notes.append(f"Only {filled}/{total_locked} LOCKED components have marks entered.")

            att_ready   += 1 if a_ready else 0
            marks_ready += 1 if m_ready else 0
            fully       += 1 if (a_ready and m_ready) else 0

            students.append(StudentReadiness(
                student_id=UUID(str(row["student_id"])),
                student_name=row["student_name"],
                usn=row.get("usn"),
                attendance_ready=a_ready,
                internal_marks_ready=m_ready,
                attendance_pct=att_pct,
                marks_fill_pct=marks_fill_pct,
                notes=notes,
            ))

        return SectionReadinessOut(
            section_id=section_id,
            section_name=section_name,
            total_students=len(students),
            attendance_ready_count=att_ready,
            internal_marks_ready_count=marks_ready,
            fully_ready_count=fully,
            students=students,
        )
