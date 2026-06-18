"""Integration tests for bulk user onboarding.

All tests run against a real PostgreSQL database and a real FastAPI app.
A fresh tenant is provisioned per module so each test run is isolated.

Covered scenarios
-----------------
1. ADMIN can bulk-generate students — USN format, email format, DB state.
2. Duplicate USN/email prevention — re-running same range skips existing rows.
3. CSV student import preview — validates rows without writing to DB.
4. CSV student import commit — inserts valid rows, reports skipped.
5. CSV faculty import — creates FACULTY role users.
6. RBAC — FACULTY cannot access any onboarding endpoint.
7. Tenant isolation — generated users are scoped to the provisioned tenant.
"""

import io
import random
import uuid
from pathlib import Path

import pytest
import pytest_asyncio
from httpx import ASGITransport, AsyncClient
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine

from app.config import settings
from app.core.auth.security import create_access_token, hash_password
from app.main import app

_BACKEND_DIR = str(Path(__file__).parent.parent.parent)

_engine = create_async_engine(
    settings.DATABASE_URL,
    echo=False,
    connect_args={"statement_cache_size": 0},
)
_Session = async_sessionmaker(_engine, class_=AsyncSession, expire_on_commit=False)

_RUN_ID = uuid.uuid4().hex[:8]
TENANT_NAME = f"Onboard Test {_RUN_ID}"
ADMIN_EMAIL = f"admin-ob-{_RUN_ID}@obtest.edu"
ADMIN_PASSWORD = "OBAdmin1234!"


# ---------------------------------------------------------------------------
# Shared fixtures
# ---------------------------------------------------------------------------

@pytest.fixture(scope="module", autouse=True)
def _setup_public(setup_public_schema):  # noqa: ARG001
    pass


@pytest_asyncio.fixture(scope="module")
async def http_client():
    ip = f"10.{random.randint(1, 254)}.{random.randint(1, 254)}.2"
    transport = ASGITransport(app=app, client=(ip, 9001))
    async with AsyncClient(transport=transport, base_url="http://test") as client:
        yield client


@pytest_asyncio.fixture(scope="module")
async def platform_user():
    uid = uuid.uuid4()
    email = f"superadmin-ob-{uid.hex[:8]}@test.com"
    pw_hash = hash_password("SuperOB1234!")
    async with _Session() as session:
        async with session.begin():
            await session.execute(
                text(
                    "INSERT INTO public.platform_users "
                    "(id, email, password_hash, full_name, is_active) "
                    "VALUES (:id, :email, :hash, :name, true)"
                ),
                {"id": str(uid), "email": email, "hash": pw_hash, "name": "Super OB"},
            )
    yield {"id": uid, "email": email}
    async with _Session() as session:
        async with session.begin():
            await session.execute(
                text("DELETE FROM public.platform_users WHERE id = :id"),
                {"id": str(uid)},
            )


def _platform_headers(user: dict) -> dict:
    token = create_access_token({
        "sub": str(user["id"]),
        "tenant_id": None,
        "schema_name": None,
        "role": "SUPER_ADMIN",
        "email": user["email"],
    })
    return {"Authorization": f"Bearer {token}"}


@pytest_asyncio.fixture(scope="module")
async def provisioned(http_client, platform_user):
    resp = await http_client.post(
        "/tenants",
        json={
            "name": TENANT_NAME,
            "admin_email": ADMIN_EMAIL,
            "admin_password": ADMIN_PASSWORD,
            "admin_full_name": "OB Admin",
        },
        headers=_platform_headers(platform_user),
    )
    assert resp.status_code == 201, f"Provisioning failed: {resp.text}"
    data = resp.json()
    yield data
    async with _Session() as session:
        async with session.begin():
            await session.execute(
                text("DELETE FROM public.refresh_token_index WHERE schema_name = :s"),
                {"s": data["schema_name"]},
            )
            await session.execute(
                text("UPDATE public.tenants SET is_active = false WHERE slug = :slug"),
                {"slug": data["slug"]},
            )
            await session.execute(text(f"DROP SCHEMA IF EXISTS {data['schema_name']} CASCADE"))


@pytest_asyncio.fixture(scope="module")
async def admin_headers(http_client, provisioned):
    slug = provisioned["slug"]
    resp = await http_client.post(
        "/auth/login",
        json={"email": ADMIN_EMAIL, "password": ADMIN_PASSWORD},
        headers={"X-Tenant-Slug": slug},
    )
    assert resp.status_code == 200
    token = resp.json()["access_token"]
    return {"Authorization": f"Bearer {token}", "X-Tenant-Slug": slug}


# ---------------------------------------------------------------------------
# Helper to count users in the tenant schema
# ---------------------------------------------------------------------------

async def _count_users(schema: str, role: str | None = None) -> int:
    async with _Session() as session:
        async with session.begin():
            await session.execute(text(f"SET LOCAL search_path = {schema}, public"))
            if role:
                result = await session.execute(
                    text("SELECT COUNT(*) FROM users WHERE role = CAST(:role AS tenantrole)"),
                    {"role": role},
                )
            else:
                result = await session.execute(text("SELECT COUNT(*) FROM users"))
            return result.scalar_one()


# ---------------------------------------------------------------------------
# Test 1 — Bulk generate students
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_generate_students(http_client, provisioned, admin_headers):
    schema = provisioned["schema_name"]
    before = await _count_users(schema, "STUDENT")

    resp = await http_client.post(
        "/admin/onboarding/generate-students",
        json={
            "usn_prefix": "OBT",
            "program_code": "MCA",
            "batch_year": 26,
            "section": "A",
            "count": 10,
            "start_seq": 1,
            "seq_width": 3,
            "email_domain": "obtest.edu",
            "default_password": "Student@123",
        },
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["created"] == 10
    assert data["skipped"] == 0
    assert data["duplicate_usns"] == []
    assert data["duplicate_emails"] == []
    assert data["default_password"] == "Student@123"

    after = await _count_users(schema, "STUDENT")
    assert after - before == 10


@pytest.mark.asyncio
async def test_generate_students_usn_format(provisioned, admin_headers, http_client):
    schema = provisioned["schema_name"]
    resp = await http_client.post(
        "/admin/onboarding/generate-students",
        json={
            "usn_prefix": "XYZ",
            "program_code": "BCE",
            "batch_year": 25,
            "count": 3,
            "start_seq": 1,
            "seq_width": 3,
            "email_domain": "obtest.edu",
        },
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text

    async with _Session() as session:
        async with session.begin():
            await session.execute(text(f"SET LOCAL search_path = {schema}, public"))
            result = await session.execute(
                text("SELECT identifier FROM users WHERE identifier LIKE 'XYZ25BCE%' ORDER BY identifier")
            )
            usns = [row[0] for row in result.fetchall()]

    assert usns == ["XYZ25BCE001", "XYZ25BCE002", "XYZ25BCE003"]


# ---------------------------------------------------------------------------
# Test 2 — Duplicate prevention
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_generate_students_duplicate_prevention(http_client, provisioned, admin_headers):
    # Generate once
    payload = {
        "usn_prefix": "DUP",
        "program_code": "CSE",
        "batch_year": 26,
        "count": 5,
        "start_seq": 1,
        "email_domain": "obtest.edu",
    }
    r1 = await http_client.post("/admin/onboarding/generate-students", json=payload, headers=admin_headers)
    assert r1.status_code == 200
    assert r1.json()["created"] == 5

    # Generate same range again — all should be skipped
    r2 = await http_client.post("/admin/onboarding/generate-students", json=payload, headers=admin_headers)
    assert r2.status_code == 200
    data = r2.json()
    assert data["created"] == 0
    assert data["skipped"] == 5
    assert len(data["duplicate_usns"]) == 5


# ---------------------------------------------------------------------------
# Test 3 — CSV student import preview (no DB writes)
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_csv_students_preview(http_client, provisioned, admin_headers):
    schema = provisioned["schema_name"]
    before = await _count_users(schema, "STUDENT")

    # program_code is now required — rows without a resolvable program are invalid
    csv_content = (
        "full_name,email,identifier,program_code\n"
        "Alice Brown,alice.brown@obtest.edu,CSV26MCA001,NOTEXIST\n"  # program not found → invalid
        "Bob Green,not-an-email,CSV26MCA002,NOTEXIST\n"             # invalid email → invalid
        "Charlie White,charlie@obtest.edu,CSV26MCA003,NOTEXIST\n"   # program not found → invalid
    )

    resp = await http_client.post(
        "/admin/onboarding/import/students/preview",
        files={"file": ("students.csv", io.BytesIO(csv_content.encode()), "text/csv")},
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["total_rows"] == 3
    assert data["valid_rows"] == 0
    assert data["invalid_rows"] == 3

    # No writes during preview
    assert await _count_users(schema, "STUDENT") == before


# ---------------------------------------------------------------------------
# Test 4 — CSV student import commit
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_csv_students_commit(http_client, provisioned, admin_headers):
    schema = provisioned["schema_name"]
    before = await _count_users(schema, "STUDENT")

    # program_code is required — rows without a resolvable program are skipped
    # (no program set up in this tenant, so all rows will be invalid)
    csv_content = (
        "full_name,email,identifier,program_code\n"
        "Dave Black,dave.black@obtest.edu,CMT26MCA001,NOTEXIST\n"
        "Eve White,eve.white@obtest.edu,CMT26MCA002,NOTEXIST\n"
    )

    resp = await http_client.post(
        "/admin/onboarding/import/students/commit",
        data={"default_password": "Student@123"},
        files={"file": ("students.csv", io.BytesIO(csv_content.encode()), "text/csv")},
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    # Both rows invalid (program not found) → 0 created, 2 skipped
    assert data["created"] == 0
    assert data["skipped"] == 2

    assert await _count_users(schema, "STUDENT") == before


# ---------------------------------------------------------------------------
# Test 5 — CSV faculty import
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_csv_faculty_commit(http_client, provisioned, admin_headers):
    schema = provisioned["schema_name"]
    before = await _count_users(schema, "FACULTY")

    csv_content = (
        "full_name,email,employee_id,department,designation\n"
        "Dr. Frank Singh,frank.singh@obtest.edu,EMP101,Computer Science,Professor\n"
        "Dr. Gita Rao,gita.rao@obtest.edu,EMP102,Mathematics,Associate Professor\n"
    )

    resp = await http_client.post(
        "/admin/onboarding/import/faculty/commit",
        data={"default_password": "Faculty@123"},
        files={"file": ("faculty.csv", io.BytesIO(csv_content.encode()), "text/csv")},
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["created"] == 2

    after = await _count_users(schema, "FACULTY")
    assert after - before == 2

    async with _Session() as session:
        async with session.begin():
            await session.execute(text(f"SET LOCAL search_path = {schema}, public"))
            result = await session.execute(
                text("SELECT role FROM users WHERE email = 'frank.singh@obtest.edu'")
            )
            role_val = result.scalar_one()
    assert str(role_val).upper() == "FACULTY"


# ---------------------------------------------------------------------------
# Test 6 — RBAC: FACULTY cannot access onboarding endpoints
# ---------------------------------------------------------------------------

@pytest_asyncio.fixture(scope="module")
async def faculty_headers(http_client, provisioned, admin_headers):
    slug = provisioned["slug"]
    faculty_email = f"faculty-rbac-{_RUN_ID}@obtest.edu"

    # Create faculty user via admin endpoint
    cr = await http_client.post(
        "/admin/users",
        json={
            "email": faculty_email,
            "password": "Faculty@123",
            "full_name": "RBAC Faculty",
            "role": "FACULTY",
        },
        headers=admin_headers,
    )
    assert cr.status_code == 201, cr.text

    lr = await http_client.post(
        "/auth/login",
        json={"email": faculty_email, "password": "Faculty@123"},
        headers={"X-Tenant-Slug": slug},
    )
    assert lr.status_code == 200
    token = lr.json()["access_token"]
    return {"Authorization": f"Bearer {token}", "X-Tenant-Slug": slug}


@pytest.mark.asyncio
async def test_faculty_cannot_generate_students(http_client, faculty_headers):
    resp = await http_client.post(
        "/admin/onboarding/generate-students",
        json={
            "usn_prefix": "BAD",
            "program_code": "MCA",
            "batch_year": 26,
            "count": 1,
            "email_domain": "obtest.edu",
        },
        headers=faculty_headers,
    )
    assert resp.status_code == 403


@pytest.mark.asyncio
async def test_faculty_cannot_import_csv(http_client, faculty_headers):
    csv_content = "full_name,email\nTest User,test@obtest.edu\n"
    resp = await http_client.post(
        "/admin/onboarding/import/students/preview",
        files={"file": ("s.csv", io.BytesIO(csv_content.encode()), "text/csv")},
        headers=faculty_headers,
    )
    assert resp.status_code == 403


# ---------------------------------------------------------------------------
# Test 7 — Tenant isolation: generated users are scoped to their schema
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_tenant_isolation(provisioned):
    schema = provisioned["schema_name"]
    # Users generated in this test module should only be in this tenant's schema.
    async with _Session() as session:
        async with session.begin():
            await session.execute(text(f"SET LOCAL search_path = {schema}, public"))
            result = await session.execute(
                text("SELECT COUNT(*) FROM users WHERE identifier LIKE 'OBT26MCA%'")
            )
            count_in_schema = result.scalar_one()

    # Should be 10 from test_generate_students
    assert count_in_schema == 10

    # Verify these users do NOT appear in the public schema
    async with _Session() as session:
        async with session.begin():
            result = await session.execute(
                text("SELECT COUNT(*) FROM public.platform_users WHERE email LIKE 'obt26mca%'")
            )
            count_in_public = result.scalar_one()
    assert count_in_public == 0


# ---------------------------------------------------------------------------
# Test 8 — Context-aware import (program_id supplied as form param)
# ---------------------------------------------------------------------------

@pytest_asyncio.fixture(scope="module")
async def acad_program_id(http_client, admin_headers):
    """Create a real department + program via the /academics/ endpoints and return the program UUID."""
    dept_resp = await http_client.post(
        "/academics/departments",
        json={"name": "Computer Science", "code": "CS"},
        headers=admin_headers,
    )
    assert dept_resp.status_code == 201, dept_resp.text
    dept_id = dept_resp.json()["id"]

    prog_resp = await http_client.post(
        "/academics/programs",
        json={
            "department_id": dept_id,
            "name": "Master of Computer Applications",
            "code": "MCA",
            "degree_type": "PG",
            "duration_years": 2,
        },
        headers=admin_headers,
    )
    assert prog_resp.status_code == 201, prog_resp.text
    return prog_resp.json()["id"]


@pytest.mark.asyncio
async def test_import_students_with_context(http_client, admin_headers, acad_program_id, provisioned):
    schema = provisioned["schema_name"]
    before = await _count_users(schema, "STUDENT")

    # CSV without program_code — program supplied via context
    csv_content = (
        "full_name,email,identifier\n"
        "Alice Context,alice.ctx@obtest.edu,CTX001\n"
        "Bob Context,bob.ctx@obtest.edu,CTX002\n"
    )

    resp = await http_client.post(
        "/admin/onboarding/import/students/preview",
        files={"file": ("students.csv", io.BytesIO(csv_content.encode()), "text/csv")},
        data={"program_id": acad_program_id},
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["total_rows"] == 2
    assert data["valid_rows"] == 2
    assert data["invalid_rows"] == 0
    assert data["duplicate_in_file"] == 0

    # Commit with context
    commit_resp = await http_client.post(
        "/admin/onboarding/import/students/commit",
        files={"file": ("students.csv", io.BytesIO(csv_content.encode()), "text/csv")},
        data={"default_password": "Student@123", "program_id": acad_program_id},
        headers=admin_headers,
    )
    assert commit_resp.status_code == 200, commit_resp.text
    assert commit_resp.json()["created"] == 2
    assert await _count_users(schema, "STUDENT") == before + 2


@pytest.mark.asyncio
async def test_import_students_xlsx(http_client, admin_headers, acad_program_id):
    import openpyxl
    import io as _io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["full_name", "email", "identifier"])
    ws.append(["Excel User 1", "xlsx1.ob@obtest.edu", "XL001"])
    ws.append(["Excel User 2", "xlsx2.ob@obtest.edu", "XL002"])
    buf = _io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    resp = await http_client.post(
        "/admin/onboarding/import/students/preview",
        files={"file": ("students.xlsx", _io.BytesIO(xlsx_bytes),
                         "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        data={"program_id": acad_program_id},
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["total_rows"] == 2
    assert data["valid_rows"] == 2
    assert data["invalid_rows"] == 0


@pytest.mark.asyncio
async def test_preview_duplicate_counts(http_client, admin_headers):
    # faculty endpoint has no program_code requirement — easiest to test dup counts
    csv_content = (
        "full_name,email,employee_id\n"
        "User A,dup.test.a@obtest.edu,DUPID001\n"
        "User B,dup.test.a@obtest.edu,DUPID002\n"  # duplicate email in file
        "User C,dup.test.c@obtest.edu,DUPID003\n"
    )
    resp = await http_client.post(
        "/admin/onboarding/import/faculty/preview",
        files={"file": ("faculty.csv", io.BytesIO(csv_content.encode()), "text/csv")},
        headers=admin_headers,
    )
    assert resp.status_code == 200, resp.text
    data = resp.json()
    assert data["total_rows"] == 3
    assert data["duplicate_in_file"] == 1   # User B duplicates User A's email
    assert data["duplicate_in_db"] == 0


# ---------------------------------------------------------------------------
# Test 9 — Sample CSV/XLSX downloads are well-formed
# ---------------------------------------------------------------------------

@pytest.mark.asyncio
async def test_sample_csv_students(http_client, admin_headers):
    resp = await http_client.get("/admin/onboarding/sample-csv/students", headers=admin_headers)
    assert resp.status_code == 200
    assert "text/csv" in resp.headers["content-type"]
    lines = resp.text.strip().splitlines()
    assert "full_name" in lines[0]
    assert "email" in lines[0]
    assert "identifier" in lines[0]
    assert len(lines) >= 2


@pytest.mark.asyncio
async def test_sample_csv_faculty(http_client, admin_headers):
    resp = await http_client.get("/admin/onboarding/sample-csv/faculty", headers=admin_headers)
    assert resp.status_code == 200
    lines = resp.text.strip().splitlines()
    # Phase 1.5 contract: personal_email is the login; employee_id no longer collected.
    assert "full_name" in lines[0]
    assert "personal_email" in lines[0]
    assert "roles" in lines[0]
    assert "employee_id" not in lines[0]


@pytest.mark.asyncio
async def test_sample_xlsx_students(http_client, admin_headers):
    resp = await http_client.get("/admin/onboarding/sample-xlsx/students", headers=admin_headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
    # Verify the xlsx bytes are a valid workbook
    import openpyxl
    import io as _io
    wb = openpyxl.load_workbook(_io.BytesIO(resp.content))
    ws = wb.active
    header = [str(c.value).lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
    assert "full_name" in header
    assert "email" in header


@pytest.mark.asyncio
async def test_sample_xlsx_faculty(http_client, admin_headers):
    resp = await http_client.get("/admin/onboarding/sample-xlsx/faculty", headers=admin_headers)
    assert resp.status_code == 200
    assert "spreadsheetml" in resp.headers["content-type"]
